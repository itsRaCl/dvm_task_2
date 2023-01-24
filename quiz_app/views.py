from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Quiz, QuizResponse, Choice, Question
from django import forms
from django.contrib.auth.decorators import login_required
from .forms import QuizConf, QuestionConf, ChoiceConf, UpdateQuestion, UpdateChoice
from django.contrib import messages
from .scripts import reset_responses, calculate_marks
from openpyxl import Workbook

# Create your views here.
def home(request):
    context = {
        "quizes": list(Quiz.objects.all()),
    }
    return render(request, "quiz_app/home.html", context=context)


@login_required
def view_quiz(request, quiz_id):
    quiz = Quiz.objects.get(pk=quiz_id)
    if quiz.quiz_password == "":
        if request.user.QuizMaster:
            isQM = True
        else:
            isQM = False
            if quiz_id in [
                x.quiz.id for x in list(request.user.quizresponse_set.all())
            ]:
                messages.warning(
                    request,
                    "You have already attempted this quiz, redirected you to your responses",
                )
                return redirect("quiz:view-quiz-response", quiz_id, 1)

        context = {
            "quiz_id": quiz.id,
            "quiz_title": quiz.quiz_title,
            "quiz_description": quiz.quiz_description,
            "quiz_master": quiz.quiz_master,
            "isQM": isQM,
        }
        return render(request, "quiz_app/start_quiz.html", context=context)
    else:
        return redirect("quiz:view-protected-quiz", quiz_id)


@login_required
def protected_quiz_view(request, quiz_id):
    quiz = Quiz.objects.get(pk=quiz_id)
    if request.user.QuizMaster:
        isQM = True
    else:
        isQM = False

    class QuizPasswd(forms.Form):
        password = forms.CharField(max_length=250, required=True)

    context = {
        "quiz_id": quiz.id,
        "quiz_title": quiz.quiz_title,
        "quiz_description": quiz.quiz_description,
        "quiz_master": quiz.quiz_master,
        "isQM": isQM,
    }

    form = QuizPasswd()
    context["form"] = form
    if request.method == "POST":
        form = QuizPasswd(request.POST)
        if form.is_valid():
            password = form.cleaned_data.get("password")
            if password == quiz.quiz_password:
                return redirect("quiz:view-question", quiz_id, 1)
            else:
                context["form"] = QuizPasswd()
                messages.warning(request, "Enter correct password")
                return render(request, "quiz_app/protected_quiz.html", context=context)
    else:
        return render(request, "quiz_app/protected_quiz.html", context=context)


@login_required
def question_view(request, quiz_id, question_no):
    if request.user.QuizMaster:
        messages.warning(
            request,
            "You are a QuizMaster, you cannot attempt a quiz, redirected to your profile",
        )
        return redirect("profile", request.user.get_username())
    else:
        quiz = Quiz.objects.get(pk=quiz_id)
        question = list(quiz.question_set.all())[question_no - 1]
        choices = list(question.choice_set.all())
        next_question = question_no + 1
        tot_questions = len(list(quiz.question_set.all()))
        islastquestion = tot_questions == question_no
        if quiz in [x.quiz for x in list(request.user.quizresponse_set.all())]:
            quiz_response = request.user.quizresponse_set.get(quiz=quiz)
        else:
            quiz_response = QuizResponse(quiz=quiz, quiz_taker=request.user)
            quiz_response.save()
        context = {
            "question_no": question_no,
            "quiz_title": quiz.quiz_title,
            "tot_questions": tot_questions,
            "next_question": next_question,
            "question": question,
            "choices": choices,
            "islastquestion": islastquestion,
        }
        if question.type == "SCQ" or question.type == "TF":

            class Options(forms.Form):
                answer = forms.CharField(
                    required=False,
                    widget=forms.RadioSelect(
                        choices=[(x.choice_text, x.choice_text) for x in choices]
                    ),
                )

            form = Options()
            context["form"] = form
            if request.method == "POST":
                form = Options(request.POST)
                if form.is_valid():
                    selected = form.cleaned_data.get("answer")
                    if selected == "":
                        selected = None
                        status = "UA"
                        quiz_response.questionresponse_set.create(
                            question=question, choice="", status=status
                        )
                    else:
                        if question.answer.strip().lower() == selected.strip().lower():
                            status = "C"
                        else:
                            status = "IC"

                        quiz_response.questionresponse_set.create(
                            question=question,
                            choice=selected,
                            status=status,
                        )
                        if not islastquestion:
                            messages.success(request, "Question Saved successfully")
                            return redirect(f"/take/{quiz_id}/{next_question}")
                        elif islastquestion:
                            calculate_marks(quiz_response)
                            messages.success(request, "Quiz submitted successfully")
                            return redirect("quiz:quiz-home")
            else:
                return render(request, "quiz_app/view_question.html", context=context)
        elif question.type == "MCQ":

            class Options(forms.Form):
                answer = forms.MultipleChoiceField(
                    required=False,
                    widget=forms.CheckboxSelectMultiple,
                    choices=[(x.choice_text, x.choice_text) for x in choices],
                )

            form = Options()
            context["form"] = form
            if request.method == "POST":
                form = Options(request.POST)
                if form.is_valid():
                    selected = form.cleaned_data.get("answer")
                    if selected == []:
                        selected = None
                        status = "UA"
                        mcqresponse = quiz_response.mcqresponse_set.create(
                            question=question, status=status
                        )
                    else:
                        if set(
                            [x.strip().lower() for x in question.answer.split(", ")]
                        ) == set([x.strip().lower() for x in selected]):
                            status = "C"
                        elif set([x.strip().lower() for x in selected]).issubset(
                            set(
                                [x.strip().lower() for x in question.answer.split(", ")]
                            )
                        ):
                            status = "PC"
                        else:
                            status = "IC"

                        mcqresponse = quiz_response.mcqresponse_set.create(
                            question=question, status=status
                        )
                        for i in selected:
                            mcqresponse.mcqchoiceresponse_set.create(
                                choice=i,
                            )
                        if not islastquestion:
                            messages.success(request, "Question Saved successfully")
                            return redirect(f"/take/{quiz_id}/{next_question}")
                        elif islastquestion:
                            calculate_marks(quiz_response)
                            messages.success(request, "Quiz submitted successfully")
                            return redirect("quiz:quiz-home")

            else:
                return render(request, "quiz_app/view_question.html", context=context)
        elif question.type == "INT":

            class Answer(forms.Form):
                answer = forms.IntegerField()

            form = Answer()
            context["form"] = form
            if request.method == "POST":
                form = Answer(request.POST)
                if form.is_valid():
                    selected = str(form.cleaned_data.get("answer"))
                    if selected == "":
                        selected = None
                        status = "UA"
                        quiz_response.questionresponse_set.create(
                            question=question, choice="", status=status
                        )
                    else:
                        if question.answer == selected:
                            status = "C"
                        else:
                            status = "IC"

                        quiz_response.questionresponse_set.create(
                            question=question,
                            choice=selected,
                            status=status,
                        )
                        if not islastquestion:
                            messages.success(request, "Question Saved successfully")
                            return redirect(f"/take/{quiz_id}/{next_question}")
                        elif islastquestion:
                            calculate_marks(quiz_response)
                            messages.success(request, "Quiz submitted successfully")

                            return redirect("quiz:quiz-home")
            else:
                return render(request, "quiz_app/view_question.html", context=context)


@login_required
def view_quiz_response(request, quiz_id, question_number):
    user = request.user
    if quiz_id not in [x.quiz.id for x in list(user.quizresponse_set.all())]:
        messages.warning(request, "You have not given that quiz")
        return redirect("quiz:quiz-home")
    else:
        quiz_response = user.quizresponse_set.get(quiz=Quiz.objects.get(pk=quiz_id))
        question = list(quiz_response.quiz.question_set.all())[question_number - 1]
        if question.type == "SCQ" or question.type == "TF" or question.type == "INT":
            response = quiz_response.questionresponse_set.get(question=question).choice
            status = quiz_response.questionresponse_set.get(question=question).status
        elif question.type == "MCQ":
            response = list(
                quiz_response.mcqresponse_set.get(
                    question=question
                ).mcqchoiceresponse_set.all()
            )
            status = quiz_response.mcqresponse_set.get(question=question).status
        tot_questions = len(list(Quiz.objects.get(pk=quiz_id).question_set.all()))
        islastquestion = tot_questions == question_number
        correct_choice = question.answer
        context = {
            "question_number": question_number,
            "question": question,
            "response": response,
            "status": status,
            "correct_choice": correct_choice,
            "quiz_id": quiz_id,
            "next_question": question_number + 1,
            "last_question": question_number - 1,
            "islastquestion": islastquestion,
        }
        return render(request, "quiz_app/view_responses.html", context=context)


@login_required
def delete_quiz(request, quiz_id):
    if request.user.QuizMaster:
        quiz = Quiz.objects.get(pk=quiz_id)
        if request.method == "POST":
            if quiz.quiz_master == request.user:
                quiz.delete()
                messages.success(request, "The quiz has been deleted successfully")
                return redirect("quiz:quiz-home")
            else:
                messages.warning(
                    request,
                    "Quiz not deleted! You are the not the quizmaster for the quiz",
                )
                return redirect("quiz:quiz-home")
        else:
            return render(request, "quiz_app/delete_quiz.html", context={"quiz": quiz})
    else:
        messages.warning(request, "You are not a quiz master cannot delete quiz")
        return redirect("quiz:quiz-home")


@login_required
def create_quiz(request):
    if request.user.QuizMaster:
        if request.method == "POST":
            form = QuizConf(request.POST)
            print("here_2")
            if form.is_valid():
                form = form.cleaned_data
                no_of_questions = form.get("no_of_questions")
                quiz = Quiz(
                    quiz_title=form.get("quiz_title"),
                    quiz_description=form.get("quiz_description"),
                    quiz_master=request.user,
                    quiz_password=form.get("quiz_password"),
                )
                quiz.save()
                quiz_id = Quiz.objects.filter(quiz_master=request.user).last().id
                request.session["call_type"] = "CREATE"
                return redirect("quiz:create-questions", quiz_id, no_of_questions, 1)
        else:
            form = QuizConf()
            return render(request, "quiz_app/create_quiz.html", {"form": form})
    else:
        messages.warning(request, "You cannot create a quiz, not a quiz master")
        return redirect("quiz:quiz-home")


@login_required
def add_questions(request, quiz_id, total_questions, current_question):
    quiz = Quiz.objects.get(pk=quiz_id)
    if request.method == "POST":
        form = QuestionConf(request.POST)
        if form.is_valid():
            form = form.cleaned_data
            no_of_choices = form.get("no_of_options")
            question = Question(
                question_text=form.get("question_text"),
                quiz=quiz,
                c_marks=form.get("c_marks"),
                ic_marks=form.get("ic_marks"),
                type=form.get("type"),
                answer=form.get("answer"),
            )
            question.save()
            if form.get("type") == "SCQ" or form.get("type") == "MCQ":
                if no_of_choices > 0:
                    return redirect(
                        "quiz:create-choice",
                        quiz_id,
                        total_questions,
                        current_question,
                        no_of_choices,
                        1,
                    )
                else:
                    messages.warning(request, "Please enter number of choices")
                    return redirect(
                        "quiz:create-questions",
                        quiz_id,
                        total_questions,
                        current_question,
                    )
            elif form.get("type") == "TF":
                choice = Choice(
                    choice_text="True",
                    question=quiz.question_set.all().last(),
                )
                choice.save()
                choice = Choice(
                    choice_text="False",
                    question=quiz.question_set.all().last(),
                )
                choice.save()
                if request.session["call_type"] == "CREATE":
                    messages.success(request, "Quiz has been added successfully")
                    return redirect("quiz:quiz-home")
                elif request.session["call_type"] == "UPDATE":
                    reset_responses(quiz_id)
                    messages.success(request, "Quiz has been updated successfully")
                    return redirect("quiz:update-quiz", quiz_id)
            elif form.get("type") == "INT":
                if request.session["call_type"] == "CREATE":
                    messages.success(request, "Quiz has been added successfully")
                    return redirect("quiz:quiz-home")
                elif request.session["call_type"] == "UPDATE":
                    reset_responses(quiz_id)
                    messages.success(request, "Quiz has been updated successfully")
                    return redirect("quiz:update-quiz", quiz_id)
            else:
                return redirect(
                    "quiz:create-question",
                    quiz_id,
                    total_questions,
                    current_question,
                )
    else:
        form = QuestionConf()
        return render(
            request,
            "quiz_app/create_question.html",
            {"form": form},
        )


def add_choice(
    request, quiz_id, total_questions, current_question, total_choices, current_choice
):
    if request.method == "POST":
        quiz = Quiz.objects.get(pk=quiz_id)
        form = ChoiceConf(request.POST)
        if form.is_valid():
            form = form.cleaned_data
            choice = Choice(
                choice_text=form.get("choice_text"),
                question=quiz.question_set.all().last(),
            )
            choice.save()
            if current_choice == total_choices:
                current_question += 1

                if current_question > total_questions:
                    if request.session["call_type"] == "CREATE":
                        messages.success(request, "Quiz has been added successfully")
                        return redirect("quiz:quiz-home")
                    elif request.session["call_type"] == "UPDATE":
                        reset_responses(quiz_id)
                        messages.success(request, "Quiz has been updated successfully")
                        return redirect("quiz:update-quiz", quiz_id)
                else:
                    return redirect(
                        "quiz:create-questions",
                        quiz_id,
                        total_questions,
                        current_question,
                    )
            else:
                current_choice += 1
                return redirect(
                    "quiz:create-choice",
                    quiz_id,
                    total_questions,
                    current_question,
                    total_choices,
                    current_choice,
                )
    else:
        form = ChoiceConf()
        return render(
            request,
            "quiz_app/create_choice.html",
            {"form": form},
        )


@login_required
def update_quiz(request, quiz_id):
    quiz = Quiz.objects.get(pk=quiz_id)
    if request.user.QuizMaster and request.user == quiz.quiz_master:
        request.session["call_type"] = "UPDATE"
        return render(
            request,
            "quiz_app/update_quiz.html",
            {"quiz": quiz, "questions": list(quiz.question_set.all())},
        )
    else:
        messages.warning(
            request,
            "You are not a quiz master cannot update quiz or you are no the quiz master for the quiz",
        )
        return redirect("quiz:quiz-home")


@login_required
def delete_question(request, quiz_id, question_number):
    quiz = Quiz.objects.get(pk=quiz_id)
    if request.user.QuizMaster and request.user == quiz.quiz_master:
        question = list(quiz.question_set.all())[question_number - 1]
        if request.method == "POST":
            reset_responses(quiz_id)
            question.delete()
            messages.success(request, "The question has been deleted successfully")
            return redirect("quiz:update-quiz", quiz_id)
        else:
            return render(
                request,
                "quiz_app/delete_question.html",
                {
                    "quiz_id": quiz_id,
                },
            )
    else:
        messages.warning(
            "You are not a quiz master cannot update quiz or you are no the quiz master for the quiz"
        )


@login_required
def delete_choice(request, quiz_id, question_number, choice_number):
    quiz = Quiz.objects.get(pk=quiz_id)
    question = list(quiz.question_set.all())[question_number - 1]
    if request.user.QuizMaster and request.user == quiz.quiz_master:
        choice = list(question.choice_set.all())[choice_number - 1]
        if request.method == "POST":
            choice.delete()
            reset_responses(quiz_id)
            messages.success(request, "The choice has been deleted successfully")
            return redirect("quiz:update-quiz", quiz_id)
        else:
            return render(
                request,
                "quiz_app/delete_choice.html",
                {
                    "quiz_id": quiz_id,
                },
            )
    else:
        messages.warning(
            "You are not a quiz master cannot update quiz or you are no the quiz master for the quiz"
        )


@login_required
def update_question(request, quiz_id, question_number):
    quiz = Quiz.objects.get(pk=quiz_id)
    question = list(quiz.question_set.all())[question_number - 1]
    question_dict = {
        "question_text": question.question_text,
        "c_marks": question.c_marks,
        "ic_marks": question.ic_marks,
        "answer": question.answer,
    }
    if request.user.QuizMaster and request.user == quiz.quiz_master:
        if request.method == "POST":
            form = UpdateQuestion(request.POST)
            if form.is_valid():
                form = form.cleaned_data
                question.question_text = form.get("question_text")
                question.c_marks = form.get("c_marks")
                question.ic_marks = form.get("ic_marks")
                question.answer = form.get("answer")
                question.save()
                reset_responses(quiz_id)
                messages.success(request, "The question has been updated successfully")
                return redirect("quiz:update-quiz", quiz_id)
        else:
            form = UpdateQuestion(question_dict)
            return render(
                request,
                "quiz_app/update_question.html",
                {"form": form, "question": question},
            )
    else:
        messages.warning(
            "You are not a quiz master cannot update quiz or you are no the quiz master for the quiz"
        )


@login_required
def update_choice(request, quiz_id, question_number, choice_number):
    quiz = Quiz.objects.get(pk=quiz_id)
    question = list(quiz.question_set.all())[question_number - 1]
    choice = list(question.choice_set.all())[choice_number - 1]
    if request.user.QuizMaster and request.user == quiz.quiz_master:
        if request.method == "POST":
            form = UpdateChoice(request.POST)
            if form.is_valid():
                form = form.cleaned_data
                choice.choice_text = form.get("choice_text")
                choice.save()
                reset_responses(quiz_id)
                messages.success(request, "The choice has been updated successfully")
                return redirect("quiz:update-quiz", quiz_id)
        else:
            form = UpdateChoice({"choice_text": choice.choice_text})
            return render(
                request,
                "quiz_app/update_choice.html",
                {"form": form, "choice": choice},
            )
    else:
        messages.warning(
            "You are not a quiz master cannot update quiz or you are no the quiz master for the quiz"
        )


def view_leaderboard(request, quiz_id):
    quiz = Quiz.objects.get(pk=quiz_id)
    quiz_leaders = quiz.quizresponse_set.order_by("-marks_secured")[:10]
    context = {"leaders": quiz_leaders}
    return render(request, "quiz_app/leadboard.html", context=context)


@login_required
def export_quiz(request, quiz_id):
    quiz = Quiz.objects.get(pk=quiz_id)
    if quiz.quiz_master == request.user:
        quiz_responses = list(quiz.quizresponse_set.all())
        response = HttpResponse(content_type="application/ms-excel")
        filename = f"{quiz.id}_{quiz.quiz_title}_responses.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Responses"
        ws.cell(row=1, column=1).value = "SrNo."
        ws.cell(row=1, column=2).value = "Username"
        ws.cell(row=1, column=3).value = "Email"
        ws.cell(row=1, column=4).value = "Marks Secured"
        for i in range(len(quiz_responses)):
            quiz_response = quiz_responses[i]
            ws.cell(row=i + 2, column=1).value = i + 1
            ws.cell(row=i + 2, column=2).value = quiz_response.quiz_taker.username
            ws.cell(row=i + 2, column=3).value = quiz_response.quiz_taker.email
            ws.cell(row=i + 2, column=4).value = quiz_response.marks_secured
        wb.save(response)
        return response
    else:
        messages.warning(
            request,
            "You are not the quiz master for this quiz you cannot access this link",
        )
        return redirect("quiz:quiz-home")
